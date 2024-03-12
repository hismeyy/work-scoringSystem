import os
import tkinter as tk
from datetime import datetime
from tkinter import filedialog

import openpyxl
import webview
from anytree import *

from script.get_config import get_config
from script.parsing_excel import read_excel_file, get_operation, get_project, get_score

operation_start_row = None  # 操作开始行
score_start = None  # 分数开始行
project_end = None  # 打分项结束行
column = None  # 列数

g_operation = []  # 操作
g_score = []  # 分数
g_project = []  # 项目

g_score_tree_objs = []  # 转换成树的结构后的对于关系

# 使用字典存储节点，以确保唯一性
nodes = {}

current_index = 0  # 当前的操作步骤 0是操作层 1以后就是打分层也就是进了树的第一层

g_car_model = ""  # 车模型
g_tester_name = ""  # 测试员姓名
g_file_path = ""  # 文件


def update_leaf_nodes_to_six(node):
    """更新给定节点下所有值为None的末端节点的值为6"""
    # 检查是否为末端节点
    if not node.children:  # 如果没有子节点，即为末端节点
        if node.name is None:  # 检查值是否为None
            node.name = 6  # 更新值为6
    else:
        # 如果不是末端节点，递归检查其子节点
        for child in node.children:
            update_leaf_nodes_to_six(child)


def set_leaf_values_to_six(root, target_node_name):
    """找到具有给定名称的节点，并更新其所有末端子节点的值"""
    target_node = find_by_attr(root, name="name", value=target_node_name)
    if target_node:
        update_leaf_nodes_to_six(target_node)
    else:
        print(f"未找到名为 '{target_node_name}' 的节点")


# 创建节点的函数
def create_node(name, parent):
    global nodes
    if name not in nodes:
        nodes[name] = Node(name, parent=parent)
    return nodes[name]


def get_node(name):
    global g_project
    # 构建树
    # 创建根节点
    root = Node(name)
    # 在构建树的循环中，检查是否到达每个分支的末端，并添加空节点
    for i in range(len(g_project[0])):
        parent = root
        for j in range(len(g_project)):
            node_name = g_project[j][i]
            if node_name not in nodes:
                nodes[node_name] = Node(node_name, parent=parent)
            else:
                # 如果节点已存在且不是当前父节点的子节点，则创建一个新的节点
                if nodes[node_name].parent != parent:
                    nodes[node_name] = Node(node_name, parent=parent)
            parent = nodes[node_name]

            # 如果到达了当前分支的末端
            if j == len(g_project) - 1:
                # 直接添加一个值为 None 的子节点
                Node(None, parent=parent)
    return root


def delete_nodes_based_on_array(root, keep_array):
    # Step 1: Find all leaf nodes
    leaf_nodes = [node for node in root.leaves]

    # Step 2: Delete nodes based on the array
    for i, keep in enumerate(keep_array):
        if not keep:  # If the array indicates to delete the node
            node_to_delete = leaf_nodes[i]
            while node_to_delete:
                parent = node_to_delete.parent
                node_to_delete.parent = None  # Remove the node from the tree
                # Check if the parent should also be deleted
                if parent and not parent.children:
                    node_to_delete = parent
                else:
                    break


def check_for_none_leaf(root):
    # 使用PreOrderIter进行树的前序遍历
    for node in PreOrderIter(root):
        # 检查是否为叶节点（没有子节点）
        if not node.children:  # 如果节点没有子节点
            if node.name is None:  # 如果节点的名称是 None
                return True
    return False


def update_none_leaf_to_six(root):
    for node in PreOrderIter(root):
        # 检查是否为叶节点（没有子节点）
        if not node.children:  # 如果节点没有子节点
            if node.name is None:  # 如果节点的名称是 None
                node.name = 6  # 将节点的名称改为 6


def get_layer_data(root, target_level):
    # 当前层级的节点列表，开始时只包含根节点
    current_level_nodes = [root]

    # 当前层级，开始于根节点的层级
    current_level = 0

    # 当我们还没到达目标层级时，继续遍历
    while current_level < target_level:
        # 准备下一层的节点列表
        next_level_nodes = []
        # 遍历当前层级的每个节点，收集它们的子节点
        for node in current_level_nodes:
            next_level_nodes.extend(node.children)
        # 准备下一次循环，当前层级的节点变为下一层的节点
        current_level_nodes = next_level_nodes
        # 层级增加
        current_level += 1

    # 当循环结束时，current_level_nodes 包含目标层级的所有节点
    # 获取并返回这些节点的数据（这里假设数据存储在节点的 name 属性中）
    return [node.name for node in current_level_nodes]


def has_none_leaf(node):
    """检查给定节点下是否有末端节点是None"""
    if not node.children:  # 如果是叶节点
        return node.name is None  # 检查名称是否为None
    return any(has_none_leaf(child) for child in node.children)  # 递归检查所有子节点


def find_node_and_check_none(root, target_name):
    """找到特定名称的节点，并检查其末端子节点是否包含None"""
    target_node = find(root, lambda node: node.name == target_name)  # 查找具有指定名称的节点
    if target_node:
        return has_none_leaf(target_node)  # 检查找到的节点的末端子节点是否有None
    else:
        return False  # 如果没有找到节点，返回False


def get_children_of_node_by_name(root, target_name):
    """根据节点名称返回其下一层子节点"""
    target_node = find_by_attr(root, name="name", value=target_name)
    if target_node:
        return [child.name for child in target_node.children]
    else:
        return []


def set_leaf_nodes_value(root, target_node_name, new_value):
    """设置指定节点的所有末端子节点的值"""
    target_node = find_by_attr(root, name="name", value=target_node_name)
    if target_node:
        for child in target_node.children:
            # 假设我们用节点的 'name' 属性来存储值
            child.name = new_value
    else:
        print(f"未找到名为 '{target_node_name}' 的节点")


def get_parent_node_name(root, target_node_name):
    """根据节点名称返回其父节点的名称"""
    target_node = find_by_attr(root, name="name", value=target_node_name)
    if target_node and target_node.parent:
        return target_node.parent.name
    else:
        return None


def get_leaf_values(root):
    """提取并返回树中所有叶子节点的值的列表"""
    # 使用 findall 查找所有叶子节点
    leaf_nodes = findall(root, filter_=lambda node: not node.children)
    # 从这些叶子节点中提取值并放入列表中
    leaf_values = [node.name for node in leaf_nodes]
    return leaf_values


class Api:
    def get_file(self):
        """
        获取文件路径
        """
        root = tk.Tk()
        root.withdraw()
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        root.quit()
        root.destroy()
        return file_path

    def start(self, file_path, car_model, tester_name, flag):
        global g_operation
        global g_project
        global g_score
        global g_score_tree_objs
        global current_index
        global g_car_model
        global g_tester_name
        global g_file_path
        if flag:
            g_operation = []  # 操作
            g_score = []  # 分数
            g_project = []  # 项目

            g_score_tree_objs = []  # 转换成树的结构后的对于关系

            g_car_model = car_model
            g_tester_name = tester_name
            g_file_path = file_path

            all_data = read_excel_file(file_path)
            g_operation = get_operation(all_data, operation_start_row)
            g_project = get_project(all_data, project_end)
            g_score = get_score(all_data, score_start)

            i = operation_start_row
            for operation in g_operation:
                root = get_node(operation[0])

                delete_nodes_based_on_array(root, operation[1:])
                obj = {
                    "index": i,
                    "operation": operation[0],
                    "root": root
                }
                g_score_tree_objs.append(obj)
                i = i + 1

        current_index = 0
        return_data = []
        for obj in g_score_tree_objs:
            flag = True
            root = obj["root"]
            if check_for_none_leaf(root):
                flag = False
            return_data.append({
                "index": obj["index"],
                "operation": obj["operation"],
                "flag": flag
            })
        return return_data

    def default_score(self, index):
        global g_score_tree_objs
        for obj in g_score_tree_objs:
            if obj["index"] == int(index):
                update_none_leaf_to_six(obj["root"])
                return

    def next_page(self, index):
        global current_index
        # index 是操作的所有用来获取树
        global g_score_tree_objs
        for obj in g_score_tree_objs:
            if obj["index"] == int(index):
                current_index = current_index + 1
                # 拿到当前操作 下的第一层数据
                layer_list = get_layer_data(obj["root"], current_index)
                return_data = []
                for layer in layer_list:
                    flag = True
                    if find_node_and_check_none(obj["root"], layer):
                        flag = False
                    return_data.append({
                        "project": layer,
                        "flag": flag
                    })
                print(f"下一页{current_index}")
                return return_data

    def get_other_project(self, index, project_name):
        global current_index
        if current_index < project_end:
            # index 是操作的所有用来获取树
            global g_score_tree_objs
            for obj in g_score_tree_objs:
                if obj["index"] == int(index):
                    current_index = current_index + 1
                    child_root = get_children_of_node_by_name(obj["root"], project_name)
                    return_data = []
                    for layer in child_root:
                        flag = True
                        if find_node_and_check_none(obj["root"], layer):
                            flag = False
                        return_data.append({
                            "project": layer,
                            "flag": flag
                        })
                    print(f"其他页{current_index}")
                    return return_data
        return "score"

    def get_score_data(self, index, project_name):
        # 通过project_name去比对 我要拿第几列的分数
        global g_score
        global g_project
        max_project = g_project[project_end - 1]
        i = 0  # 看是第几列
        for project in max_project:
            if project == project_name:
                # 说明i就是当前需要的列
                return_data = []
                for score in g_score:
                    return_data.append(score[i + 1])
                return return_data
            i = i + 1

    def set_score(self, index, project_name, score):
        global current_index
        # index 是操作的所有用来获取树
        global g_score_tree_objs
        for obj in g_score_tree_objs:
            if obj["index"] == int(index):
                current_index = current_index - 1
                # 找到了需要修改分数的地方
                set_leaf_nodes_value(obj["root"], project_name, score)
                parent_name = get_parent_node_name(obj["root"], project_name)
                print(f"设置分{current_index}")
                return parent_name

    def up_page(self, index, project_name):
        global current_index
        # index 是操作的所有用来获取树
        global g_score_tree_objs

        if current_index > 2:
            for obj in g_score_tree_objs:
                if obj["index"] == int(index):
                    set_leaf_values_to_six(obj["root"], project_name)
                    current_index = current_index - 2
                    parent_name = get_parent_node_name(obj["root"], project_name)
                    print(f"上一页{current_index}")
                    return parent_name

        current_index = current_index - 2

        if current_index < 0:
            return "goto0"

        for obj in g_score_tree_objs:
            if obj["index"] == int(index):
                set_leaf_values_to_six(obj["root"], project_name)
        return "goto1"

    def out_score(self):
        global g_score_tree_objs
        global g_car_model
        global g_tester_name
        global g_file_path
        root = tk.Tk()
        root.withdraw()  # 隐藏主窗口
        folder_path = filedialog.askdirectory()  # 弹出对话框让用户选择文件夹
        root.quit()
        root.destroy()

        # 获取当前日期
        current_date = datetime.now().strftime("%Y.%m.%d")

        # 构建文件名
        file_name = f"{g_car_model}-{g_tester_name}-{current_date}.xlsx"

        result_score = []

        for obj in g_score_tree_objs:
            root = obj["root"]
            value = get_leaf_values(root)
            result_score.append(value)

        print(result_score)

        workbook = openpyxl.load_workbook(g_file_path)
        sheet = workbook.active

        global operation_start_row  # 17
        global column  # 20
        max_row = len(g_operation) + operation_start_row

        for temp_row in range(operation_start_row, max_row):
            for temp_column in range(column, 1, -1):
                value = sheet.cell(row=temp_row, column=temp_column).value
                if int(value) == 1:
                    sheet.cell(row=temp_row, column=temp_column).value = result_score[
                        temp_row - operation_start_row].pop()
                else:
                    sheet.cell(row=temp_row, column=temp_column).value = None

        sheet['A1'] = g_car_model
        sheet['A2'] = os.path.basename(g_file_path)
        sheet['A3'] = g_tester_name
        output_path = os.path.join(folder_path, file_name)
        # 保存工作簿
        workbook.save(output_path)


if __name__ == '__main__':
    # 初始化API和webview窗口
    api = Api()

    # 获取当前脚本所在目录的路径
current_dir = os.path.dirname(os.path.abspath(__file__))
# 构建HTML文件的路径
html_file_path = os.path.join(current_dir, './views/new_v/index.html')

# 使用file协议和绝对路径加载本地HTML文件
url = 'file://' + html_file_path

window = webview.create_window(
    title='车型测试程序',
    url=url,  # 修改这里来加载指定的HTML文件
    js_api=api

)

# 获取配置
operation_start_row, score_start, project_end, column = get_config()
# webview.start(debug=True)
webview.start()
